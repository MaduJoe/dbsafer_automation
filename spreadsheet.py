import os
import sys
import shared
import gspread
import my_utils
import pandas as pd
import tkinter as tk

from oauth2client.service_account import ServiceAccountCredentials
from google.oauth2.service_account import Credentials
from googleapiclient.http import MediaFileUpload
from googleapiclient.discovery import build
from openpyxl import load_workbook
from tkinter import messagebox
from datetime import datetime
import my_utils


class SpreadSheets:
    def __init__(self):
        data = my_utils.read_config(shared.CONFIG)
        meta_sheet = data['google_sheet']
        self.credential_file = meta_sheet['secret_key']
        self.scope = meta_sheet['sheet_scope']
        self.sheet_id = meta_sheet['sheet_id']
        self.sheet_url = meta_sheet['sheet_url']
        self.save_dir = os.path.join(data['log']['base_path'], data['log']['dir_name'][3]) # 3 : xlsxs
    
    def convert_fail_to_json(self):
        my_utils.print_message("Save Failed JSON File")
        # 파일 가져오기 
        xlsx_file = my_utils.get_xlsx()
        wb = load_workbook(xlsx_file)
        
        # 가장 마지막 생성된 시트 선택
        last_sheet_name = wb.sheetnames[-1]
        # 시트 읽기
        df = pd.read_excel(xlsx_file, sheet_name=last_sheet_name, engine='openpyxl')
        # 데이터 영역지정 - HEADER에 해당하는 영역
        fail_rows = df[df['result'] == 'Fail'][shared.SHEET_HEADER] # 사전에 정의해 놓은 시트 해더 == SHEET_HEADER   
        # JSON 형식으로 변환
        fail_data = fail_rows.to_dict(orient='records')
        # _fail 이름 붙여서 저장 
        json_path = os.path.splitext(xlsx_file)[0].replace('xlsxs','jsons') + '_fail.json'
        my_utils.write_json(json_path, fail_data)
        
        return json_path
    
    def convert_xlsx_to_json(self, xlsx_file):
        my_utils.print_message("Convert XLSX to JSON")
        
        # Read the Excel file
        df = pd.read_excel(xlsx_file, engine='openpyxl')

        # 빈 행 삭제 (모든 컬럼이 NaN인 행 삭제)
        df = df.dropna(how='all')

        # A컬럼과 H컬럼에 숫자가 들어가있다면 float형이 아닌 int형으로 변환
        def convert_to_int_if_possible(value):
            if pd.notna(value) and isinstance(value, (int, float)) and value == int(value):
                return int(value)  # int로 변환 가능한 경우 int로 변환
            return value

        # A컬럼과 H컬럼에 대해 변환 적용
        for col in ['A', 'H']:
            if col in df.columns:
                df[col] = df[col].apply(convert_to_int_if_possible)
        
        # 모든 컬럼의 타입을 확인하여, float형이지만 소수점이 없는 경우 int로 변환
        for col in df.select_dtypes(include=['float']).columns:
            df[col] = df[col].apply(lambda x: int(x) if pd.notna(x) and x.is_integer() else x)
            

        # Convert the DataFrame to a dictionary
        data = df.to_dict(orient='records') # orient='records' : 각 행을 하나의 사전 객체로 변환함
        
        # Define the output JSON file name
        json_path = os.path.splitext(xlsx_file)[0].replace('xlsxs','jsons') + '.json'
        
        my_utils.write_json(json_path, data)
        return json_path
      
    def save_sheet(self):
        result = False
        try:
            # Google Sheets API 인증 설정
            creds = ServiceAccountCredentials.from_json_keyfile_name(self.credential_file, self.scope)
            client = gspread.authorize(creds)

            # Google Sheet 열기
            spreadsheet = client.open_by_url(self.sheet_url)
            
            worksheet = spreadsheet.sheet1
            data = worksheet.get_all_values()   # 전체 데이터를 가져오기 (헤더 포함)

            # 첫 번째 행은 헤더로 가정
            # [:-1] - select 컬럼 제외
            header = data[0][:-1]
            # print(header)
            # print(data[1][0]) # 2, 1번째 값
            # print(data[1][15]) # 2, 16번째 값

            # N 열 (index 13, 0-based)의 값이 True인 행만 필터링
            # ToDo - select컬럼 없이 저장하고 싶으면 [row for row in data[1:] if row[13].lower() == 'true'][:-1] 필요할듯
            # [:-1] - select 컬럼 제외
            # filtered_data = [header] + [row[:-1] for row in data[1:] if row[13].lower() == 'true']   
            filtered_data = [header] + [row[:-1] for row in data[1:] if row[15].lower() == 'true']   # 12번째열 (L열)
            
            # DataFrame으로 변환
            df = pd.DataFrame(filtered_data)
            df.columns = df.iloc[0]  # 첫 번째 행을 헤더로 설정
            df = df[1:]  # 첫 번째 행 제외하고 데이터만 남김
            # df = pd.DataFrame(filtered_data[1:], columns=filtered_data[0])
            print(df)

            # 파일 저장 경로 및 파일명 설정
            current_time = datetime.now().strftime("%Y%m%d_%H%M%S")
            file_name = f"manager_{current_time}.xlsx"
            file_path = os.path.join(self.save_dir, file_name)
            df.to_excel(file_path, index=False, engine='openpyxl')
            print(f"파일이 {file_path}에 저장되었습니다.")
            result = file_path
            
        except gspread.exceptions.APIError as e:
            print(f"Google Sheets API 오류가 발생했습니다: {e}")
        except FileNotFoundError as e:
            print(f"지정된 경로에 접근할 수 없습니다: {e}")
        except PermissionError as e:
            print(f"파일을 저장할 수 있는 권한이 없습니다: {e}")
        except Exception as e:
            print(f"예상치 못한 오류가 발생했습니다: {e}")
            
        finally:
            if not result:
                sys.exit(1)
            return result
    
    def ask_user_confirmation(self):
        root = tk.Tk()
        root.withdraw()  # Tkinter 창을 숨깁니다.
        result = messagebox.askquestion("확인", "구글 시트에 업데이트 하겠습니까?")
        root.destroy()  # Tkinter 창을 종료합니다.
        return result == 'yes'
    
    # def upload_or_update_sheet(self, file_path, sheet_id=None):
    #     my_utils.print_message("Update Google Sheet")
    #     if not self.ask_user_confirmation():
    #         print("업데이트가 취소되었습니다.")
    #         return
    #     try:
    #         creds = ServiceAccountCredentials.from_json_keyfile_name(self.credintial_file, self.scope)
    #         drive_service = build('drive', 'v3', credentials=creds)
    #         if sheet_id:
    #             # 기존 시트 업데이트
    #             file_metadata = {}
    #             media = MediaFileUpload(file_path, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    #             updated_file = drive_service.files().update(fileId=sheet_id, media_body=media).execute()
    #             print(f"Google Sheet가 업데이트되었습니다. 파일 ID: {updated_file.get('id')}")
    #         else:
    #             # 새로운 시트 생성
    #             file_metadata = {'name': os.path.basename(file_path), 'mimeType': 'application/vnd.google-apps.spreadsheet'}
    #             media = MediaFileUpload(file_path, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    #             uploaded_file = drive_service.files().create(body=file_metadata, media_body=media).execute()
    #             print(f"Google Sheet가 생성되었습니다. 파일 ID: {uploaded_file.get('id')}")
    #     except Exception as e:
    #         print(f"Google Sheet 업로드 중 오류가 발생했습니다: {e}")
    
    
    # ver2. 기존 데이터 덮어버리는것이 아닌 변경된 행 내용만 update되는 방식으로 변경.
    def upload_or_update_sheet(self, file_path, sheet_id=None):
        my_utils.print_message("Update Google Sheet")
        
        # if not self.ask_user_confirmation():
        #     print("업데이트가 취소되었습니다.")
        #     return
        
        try:
            # Load the local Excel file into a DataFrame
            local_df = pd.read_excel(file_path, engine='openpyxl')
            
            # Convert all data types to strings to avoid JSON serialization issues
            local_df = local_df.astype(str)
            local_df = local_df.fillna('')
            # print(f"local_df : {local_df}")
            print(f"Local Excel file columns: {local_df.shape[1]}")  # Number of columns

            # Authenticate and create the Google Sheets API service
            creds = Credentials.from_service_account_file(self.credential_file, scopes=self.scope)
            sheets_service = build('sheets', 'v4', credentials=creds)
            
            if sheet_id:
                # Get the existing Google Sheet data
                sheet = sheets_service.spreadsheets()
                # print(sheet_id)
                result = sheet.values().get(spreadsheetId=sheet_id, range='Sheet1!A:O').execute()
                google_sheet_data = result.get('values', [])
                # print(google_sheet_data[0]) # 헤더 출력
                # print(google_sheet_data[1]) # 13
                
                # Convert the Google Sheet data to a DataFrame
                google_df = pd.DataFrame(google_sheet_data[1:], columns=google_sheet_data[0])
                google_df = google_df.astype(str)
                # google_df = google_df.fillna('')
                # print(f"google_df : {google_df}")
                # print(f"Google Sheet columns: {google_df.shape[1]}")

                
                # Compare the data (excluding headers)
                changes = []
                updated_rows = set()
                for idx, local_row in local_df.iterrows():
                    if idx < len(google_df):
                        google_row = google_df.iloc[idx]
                                
                        for col_idx, col in enumerate(local_df.columns[:15]):
                            if local_row[col] != google_row[col]:
                                column_letter = chr(65 + col_idx)  # A1 표기법의 열 문자로 변환
                                range_name = f'Sheet1!{column_letter}{idx+2}'
                                changes.append({
                                    'range': range_name,
                                    'values': [[local_row[col]]]
                                })
                                updated_rows.add(idx + 2)      
                    
                    else:
                        # If there are new rows in the local file, add them to the sheet
                        new_row_range = f'Sheet1!A{idx+2}:O{idx+2}'
                        new_row_values = local_df.iloc[idx, :15].values.tolist()
                        changes.append({
                            'range': new_row_range,
                            'values': [new_row_values]
                        })
                        updated_rows.add(idx + 2)
                
                # Update the Google Sheet with the changes
                if changes:
                    body = {
                        'valueInputOption': 'RAW',
                        'data': changes
                    }
                    sheets_service.spreadsheets().values().batchUpdate(spreadsheetId=sheet_id, body=body).execute()
                    print(f"Google Sheet가 업데이트되었습니다. 업데이트된 행 수: {len(changes)}")
                    print(f"업데이트된 행 번호: {sorted(updated_rows)}")
                else:
                    print("업데이트할 내용이 없습니다.")
            
            else:
                print(f"sheet_id 가 입력되지 않았습니다 : {sheet_id}")

        except Exception as e:
            print(f"Google Sheet 업로드 중 오류가 발생했습니다: {e}")
         
        
if __name__ == '__main__':
    ss = SpreadSheets()
    # ss.save_sheet()
    
    xlsx_file = my_utils.get_xlsx()
    print(f"xlsx_file : {xlsx_file}")
    ss.upload_or_update_sheet(file_path=xlsx_file, sheet_id=ss.sheet_id)
    
    # ss.save_sheet()
    # ss.convert_fail_to_json()