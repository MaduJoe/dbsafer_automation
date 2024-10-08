import ast
import tkinter as tk
from tkinter import ttk, messagebox
import json
import yaml
from datetime import datetime
from openpyxl import load_workbook, Workbook
import os
import shared
import my_utils

class InputDialog:
    def __init__(self, excel_file=None, macro_list=""):
        my_utils.print_message("Get User Input")
        self.root = tk.Tk()
        self.root.title("Record Input Form")
        self.root.geometry("800x1200")  # Height adjusted to accommodate new fields
        self.root.resizable(False, False)
        self.root.protocol("WM_DELETE_WINDOW", self.on_close)
        self.upload_result = None

        # Center the window on the screen
        self.center_window(700, 700)
        
        # 맨 앞으로 창을 열도록 설정
        self.root.attributes("-topmost", True)

        # Load the config.yaml file
        self.load_config(shared.CONFIG)

        # Specify the Excel file name (assuming it already exists)
        self.excel_file = excel_file

        # Event list
        self.event_list = str(macro_list) if macro_list else ""

        # Create the UI
        self.create_widgets()

    def get_result(self):
        """
        This function will return the value of self.upload_result after the window is closed.
        """
        return self.upload_result
    
    def on_close(self):
        """
        This function is called when the window is closed by the user.
        It prompts the user with a 'Would you like to upload to Google Sheets?' dialog.
        If the user clicks 'Yes', it returns True.
        """
        if messagebox.askyesno("Upload Confirmation", "Would you like to upload to Google Sheets?"):
            print("User chose to upload.")
            # Perform any upload logic here (if required) before returning True.
            # self.upload_to_google_sheets()  # Call your upload function if necessary
            # self.root.quit()  # Close the tkinter window
            self.upload_result = True
            # return True
        else:
            print("User chose not to upload.")
            # self.root.quit()  # Close the tkinter window
            self.upload_result = False
            # return False
        self.root.quit()
    
    def center_window(self, width, height):
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        x = (screen_width // 2) - (width // 2)
        y = (screen_height // 2) - (height // 2)
        self.root.geometry(f'{width}x{height}+{x}+{y}')

    def load_config(self, config_path):
        with open(config_path, 'r') as file:
            self.config = yaml.safe_load(file)

    def create_widgets(self):
        # Product input field (using Entry widget to match size)
        tk.Label(self.root, text="Product:").grid(row=0, column=0, sticky=tk.W, padx=10, pady=5)
        self.product_entry = tk.Entry(self.root, width=50)
        self.product_entry.grid(row=0, column=1, padx=10, pady=5)

        # Set text and then make it readonly
        self.product_entry.insert(0, self.config['tc']['product'])  # Insert the product name
        self.product_entry.config(state="readonly")

        # Summary input field
        tk.Label(self.root, text="Summary:").grid(row=1, column=0, sticky=tk.W, padx=10, pady=5)
        self.summary_entry = tk.Entry(self.root, width=50)
        self.summary_entry.grid(row=1, column=1, padx=10, pady=5)

        tk.Label(self.root, text="Type:").grid(row=2, column=0, sticky=tk.W, padx=10, pady=5)
        self.type_var = tk.StringVar(self.root)
        self.type_dropdown = ttk.Combobox(self.root, textvariable=self.type_var, values=self.config['tc']['type'], state="readonly", width=20)
        self.type_dropdown.grid(row=2, column=1, padx=10, pady=5)
        self.type_dropdown.current(0)

        tk.Label(self.root, text="Version:").grid(row=3, column=0, sticky=tk.W, padx=10, pady=5)
        self.version_var = tk.StringVar(self.root)
        self.version_dropdown = ttk.Combobox(self.root, textvariable=self.version_var, values=self.config['tc']['version'], state="readonly", width=20)
        self.version_dropdown.grid(row=3, column=1, padx=10, pady=5)
        self.version_dropdown.current(1)

        ###
        tk.Label(self.root, text="Group:").grid(row=4, column=0, sticky=tk.W, padx=10, pady=5)
        self.group_var = tk.StringVar(self.root)
        self.group_dropdown = ttk.Combobox(self.root, textvariable=self.group_var, values=list(self.config['tc']['group'].keys()), state="readonly", width=30)
        self.group_dropdown.grid(row=4, column=1, padx=10, pady=5)
        self.group_dropdown.bind("<<ComboboxSelected>>", self.update_subgroup)
        # self.group_dropdown.current(0)

        tk.Label(self.root, text="Subgroup:").grid(row=5, column=0, sticky=tk.W, padx=10, pady=5)
        self.subgroup_var = tk.StringVar(self.root)
        self.subgroup_dropdown = ttk.Combobox(self.root, textvariable=self.subgroup_var, state="readonly", width=30)
        self.subgroup_dropdown.grid(row=5, column=1, padx=10, pady=5)
        self.subgroup_dropdown.bind("<<ComboboxSelected>>", self.update_thirdgroup)
        # self.subgroup_dropdown.current(0)

        tk.Label(self.root, text="Third Group:").grid(row=6, column=0, sticky=tk.W, padx=10, pady=5)
        self.thirdgroup_var = tk.StringVar(self.root)
        self.thirdgroup_dropdown = ttk.Combobox(self.root, textvariable=self.thirdgroup_var, state="readonly", width=30)
        self.thirdgroup_dropdown.grid(row=6, column=1, padx=10, pady=5)
        # self.thirdgroup_dropdown.current(0)
        ###
        
        tk.Label(self.root, text="Precondition:").grid(row=7, column=0, sticky=tk.W, padx=10, pady=5)
        self.precondition_var = tk.StringVar(self.root)
        self.precondition_dropdown = ttk.Combobox(self.root, textvariable=self.precondition_var, values=self.config['tc']['precondition'], state="readonly", width=20)
        self.precondition_dropdown.grid(row=7, column=1, padx=10, pady=5)
        self.precondition_dropdown.current(1)

        tk.Label(self.root, text="Macro List:").grid(row=8, column=0, sticky=tk.W, padx=10, pady=5)
        self.macro_list_entry = tk.Text(self.root, width=50, height=5)
        self.macro_list_entry.grid(row=8, column=1, padx=10, pady=5)
        self.macro_list_entry.insert(tk.END, self.event_list)

        # Macro Steps input field
        tk.Label(self.root, text="Macro Steps:").grid(row=9, column=0, sticky=tk.W, padx=10, pady=5)
        self.macro_steps_entry = tk.Text(self.root, width=50, height=5)
        self.macro_steps_entry.grid(row=9, column=1, padx=10, pady=5)

        tk.Label(self.root, text="Val Type:").grid(row=10, column=0, sticky=tk.W, padx=10, pady=5)
        self.val_type_var = tk.StringVar(self.root)
        self.val_type_dropdown = ttk.Combobox(self.root, textvariable=self.val_type_var, values=["image", "query"], state="readonly", width=20)
        self.val_type_dropdown.grid(row=10, column=1, padx=10, pady=5)
        self.val_type_dropdown.set("query")
        self.val_type_dropdown.bind("<<ComboboxSelected>>", self.on_val_type_change)

        self.dynamic_label = tk.Label(self.root, text="검증할 쿼리문을 입력하세요:")
        self.dynamic_label.grid(row=11, column=0, sticky=tk.W, padx=10, pady=5)
        self.dynamic_entry = tk.Entry(self.root, width=50)
        self.dynamic_entry.grid(row=11, column=1, padx=10, pady=5)

        # Expected Result input field
        tk.Label(self.root, text="Expected Result:").grid(row=12, column=0, sticky=tk.W, padx=10, pady=5)
        self.expected_result_entry = tk.Entry(self.root, width=50)
        self.expected_result_entry.grid(row=12, column=1, padx=10, pady=5)

        self.add_button = tk.Button(self.root, text="Add Record", command=self.add_record)
        self.add_button.grid(row=13, column=0, columnspan=2, pady=10)

    def update_subgroup(self, event):
        selected_group = self.group_var.get()
        if selected_group in self.config['tc']['group']:
            subgroup_data = self.config['tc']['group'][selected_group]
            if isinstance(subgroup_data, dict):
                self.subgroup_dropdown['values'] = list(subgroup_data.keys())
                self.subgroup_dropdown.config(state="readonly")
                self.subgroup_dropdown.set('')  # Clear the selected value
                
            else:
                self.subgroup_var.set('')  # Clear the selected value
                self.subgroup_dropdown.set('')  # Clear the selected value
                self.subgroup_dropdown['values'] = []
                self.subgroup_dropdown.config(state="disabled")
        
        self.thirdgroup_var.set('')  # Clear the selected value
        self.thirdgroup_dropdown.set('')  # Clear the selected value
        self.thirdgroup_dropdown['values'] = []
        self.thirdgroup_dropdown.config(state="disabled")

    def update_thirdgroup(self, event):
        selected_group = self.group_var.get()
        selected_subgroup = self.subgroup_var.get()
        if selected_group in self.config['tc']['group']:
            subgroup_data = self.config['tc']['group'][selected_group]
            if isinstance(subgroup_data, dict) and selected_subgroup in subgroup_data:
                third_group_data = subgroup_data[selected_subgroup]
                if third_group_data:
                    self.thirdgroup_dropdown['values'] = third_group_data
                    self.thirdgroup_dropdown.config(state="readonly")
                    self.thirdgroup_var.set('')  # Clear any previous selection in thirdgroup
                    
                else:
                    self.thirdgroup_var.set('')  # Clear the selected value
                    self.thirdgroup_dropdown.set('')    # Clear the selected value
                    self.thirdgroup_dropdown['values'] = []
                    self.thirdgroup_dropdown.config(state="disabled")
            else:
                self.thirdgroup_var.set('')  # Clear the selected value
                self.thirdgroup_dropdown.set('')    # Clear the selected value
                self.thirdgroup_dropdown['values'] = []
                self.thirdgroup_dropdown.config(state="disabled")
                
    def on_val_type_change(self, event):
        val_type = self.val_type_var.get()
        if val_type == "image":
            self.dynamic_label.config(text="검증할 이미지를 입력하세요:")
        elif val_type == "query":
            self.dynamic_label.config(text="검증할 쿼리문을 입력하세요:")
        else:
            self.dynamic_label.config(text="")
        self.dynamic_entry.delete(0, tk.END)

    def add_record(self):
        product = self.product_entry.get()
        summary = self.summary_entry.get()
        tc_type = self.type_var.get()
        version = self.version_var.get()
        group = self.group_var.get()
        sub_group = self.subgroup_var.get()
        third_group = self.thirdgroup_var.get()
        precondition = self.precondition_var.get()
        macro_list = self.macro_list_entry.get("1.0", tk.END).strip()
        macro_steps = self.macro_steps_entry.get("1.0", tk.END).strip()
        val_type = self.val_type_var.get()
        dynamic_value = self.dynamic_entry.get()
        expected_result = self.expected_result_entry.get()

        try:
            macro_list = eval(macro_list) if macro_list else []
            if val_type == "image":
                val_data = {"image_name": dynamic_value}
            elif val_type == "query":
                val_data = {"query": dynamic_value}
            else:
                val_data = {}
        except (json.JSONDecodeError, SyntaxError):
            messagebox.showerror("Error", "JSON format or list format is invalid.")
            return

        # Load the existing workbook
        if os.path.exists(self.excel_file):
            wb = load_workbook(self.excel_file)
            first_sheet_name = wb.sheetnames[0]
            print(f"first_sheet_name : {first_sheet_name}")
            ws = wb[first_sheet_name]
            ws = wb.active
            
            # Find the maximum ID in the existing data
            max_row = ws.max_row
            if max_row > 1:
                max_id_cell = ws.cell(row=max_row, column=1)
                try:
                    self.current_id = int(max_id_cell.value) + 1
                except (TypeError, ValueError):
                    self.current_id = 1
            else:
                self.current_id = 1
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "TC_LIST"
            
            headers = ["id", "product", "group1", "group2", "group3","summary", "type", "version", "precondition", "macro list", "macro steps", "val type", "val data", "expected result", "update time"]
            ws.append(headers)
            self.current_id = 1

        update_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        new_record = {
            "id": self.current_id,
            "product": product,
            "group1": group,
            "group2": sub_group,
            "group3": third_group,
            "summary": summary,
            "type": tc_type,
            "version": version,
            "precondition": precondition,
            "macro list": str(macro_list),
            "macro steps": macro_steps,
            "val type": val_type,
            "val data": json.dumps(val_data),
            "expected result": expected_result,
            "update time": update_time
        }

        self.save_to_excel(new_record, wb, ws)

        # Reset fields after successful addition
        self.summary_entry.delete(0, tk.END)
        self.macro_list_entry.delete("1.0", tk.END)
        self.macro_steps_entry.delete("1.0", tk.END)
        self.dynamic_entry.delete(0, tk.END)
        self.expected_result_entry.delete(0, tk.END)
        self.val_type_var.set("")

        messagebox.showinfo("Success", "Record added successfully!")

    def save_to_excel(self, record, wb, ws):
        # Format the macro list to be more readable in Excel
        events = ast.literal_eval(record["macro list"])  # Convert string to list
        formatted_macro_list = "[\n" + ",\n".join(str(event) for event in events) + "\n]"

        row_data = [
            record["id"],
            record["product"],
            record["group1"],
            record["group2"],
            record["group3"],
            record["summary"],
            record["type"],
            record["version"],
            record["precondition"],
            formatted_macro_list,  # Properly formatted macro list
            record["macro steps"],
            record["val type"],
            record["val data"],
            record["expected result"],
            record["update time"]
        ]
        ws.append(row_data)
        print(f"Row added: {row_data}")  # 디버깅 출력
        
        wb.save(self.excel_file)
        print(f"Saved to {self.excel_file} successfully")

# Main program execution
if __name__ == "__main__":
    event_list = [('click', ('1', 'List', 'single', 'left', (1108, 796), 'id', 1)), 
                  ('click', ('1', 'List', 'single', 'left', (977, 695), 'id', 0)), 
                  ('click', ('1', 'List', 'double', 'left', (993, 555), 'id', 0)), 
                  ('click', ('1', 'List', 'single', 'left', (1034, 783), 'id', 0)), 
                  ('keydown', 'f4', 0), 
                  ('keyup', 'f4', 0)]
    # app = InputDialog()
    # formatted_macro_list = "[\n" + ",\n".join(str(event) for event in event_list) + "\n]"


    xlsx_file = my_utils.get_xlsx()
    app = InputDialog(excel_file=xlsx_file, macro_list=event_list)  # 3.1
    app.root.mainloop()
    result = app.get_result()
    print(result)
    