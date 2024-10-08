# -*- coding: utf-8 -*-
import win32com.client
import pythoncom
import os
import sys
import yaml
import json
import time
import shared
import win32gui
import win32con
import psutil
import win32process
import glob
import my_utils

from pywinauto import Application, findwindows
from pywinauto.uia_element_info import UIAElementInfo
from pywinauto.findwindows import find_windows, ElementNotFoundError
from openpyxl.styles import PatternFill
from openpyxl import load_workbook, Workbook
import pywinauto.mouse as mo
import spreadsheet
import _ctypes
import pywinauto
import log_handler

logger = log_handler.setup_logger()

def release_services():
    try:

        # focus window
        my_utils.focus_window()        

        # 객체관리 클릭 
        x, y = (15, 140)
        mo.click(coords=(x, y), button='left')
        print(f"Clicked ({x}, {y}) Success")

        # 안펼쳐져있으면(Transparent/Bypass 항목 안보이면) 펼치기  
        ctrl_name = "Transparent/Bypass"
        ctrl_type = "TreeItem"
        element = shared.PARENT_WINDOW.child_window(title=ctrl_name, control_type=ctrl_type)  # name(title)

        if element.exists(timeout=2):
            print("서비스 하위항목들이 Release된 상태입니다")
            # pass
        else:
            x, y = (113, 107)   # 서비스 하위 항목 나열버튼 클릭 
            mo.click(coords=(x, y), button='left')
            print(f"Clicked ({x}, {y}) Success")

    except (pywinauto.findwindows.ElementNotFoundError, pywinauto.timings.TimeoutError) as e:
        print(f"Element not found or timed out: {e}")

    except _ctypes.COMError as com_error:
        print(f"COMError encountered: {com_error}")

    except Exception as e:
        print(f"An unexpected error occurred: {e}")


def click_into_group(group):
    my_utils.focus_window()
    if group=='':
        print("그룹을 지정하지 않았습니다. 프로그램을 종료합니다.")
        sys.exit(1)
        
    group_depth = group.split('_')
    print(f"Group Depth :{group_depth}")
    
    if len(group_depth)>1:
        group1 = group_depth[0]
        group2 = group_depth[1]
        group3 = None
        
        if len(group_depth)==3:
            group3 = group_depth[2]
        
        if group1 == 'policy':
            mo.click(coords=(15, 107), button='left')   # pane type이고 id값이 모두 같으므로 좌표로 하드코딩하는게 맞다 
            time.sleep(1)

            if group2 == 'dbms':
                if group3 == 'access':
                    ctrl_name = '접속제어정책'
                elif group3 == 'auth':
                    ctrl_name = '권한제어정책'
                elif group3 == 'global':
                    ctrl_name = '결과값제어정책'
                
            elif group2 == 'ftp':
                if group3 == 'access':
                    ctrl_name = '접속제어정책'
                elif group3 == 'auth':
                    ctrl_name = '권한제어정책'
            
            elif group2 == 'terminal':
                if group3 == 'access':
                    ctrl_name = '접속제어정책'
                elif group3 == 'auth':
                    ctrl_name = '권한제어정책'

            elif group2 == 'sa':
                if group3 == 'dbms':
                    ctrl_name = 'DBMS'
                elif group3 == 'ftp':
                    ctrl_name = 'FTP'
                elif group3 == 'terminal':
                    ctrl_name = 'TERMINAL'

            
        elif group1 == 'object':
            time.sleep(1)
            mo.click(coords=(15, 140), button='left')   # pane type이고 id값이 모두 같으므로 좌표로 하드코딩하는게 맞다 
            time.sleep(1)
            
            if group2 == 'service':
                time.sleep(1)
                # mo.click(coords=(152, 106), button='left') 
                # ctrl_name = '서비스'
                # time.sleep(2)
                
                if group3 == 'dbms':
                    ctrl_name = 'DBMS'
                    
                elif group3 == 'ftp':
                    ctrl_name = 'FTP'
                    
                elif group3 == 'terminal':
                    ctrl_name = 'TERMINAL'
                    
                elif group3 == 'transparent':
                    ctrl_name = 'Transparent/Bypass'
                    
            elif group2 == 'instance':
                time.sleep(0.5)
                ctrl_name = '인스턴스'
                
            elif group2 == 'ip':
                time.sleep(0.5)
                ctrl_name = 'IP 주소'
                
            elif group2 == 'access':
                time.sleep(0.5)
                ctrl_name = '접속계정'
                
            elif group2 == 'application':
                time.sleep(0.5)
                ctrl_name = '어플리케이션'
                
            elif group2 == 'cert':
                time.sleep(0.5)
                ctrl_name = '보안계정'
                
            elif group2 == 'datamask':
                time.sleep(0.5)
                ctrl_name = '데이터마스킹'
                
            elif group2 == 'tablecolumn':
                time.sleep(0.5)
                ctrl_name = '테이블/컬럼'
                
            elif group2 == 'formalquery':
                time.sleep(0.5)
                mo.click(coords=(152, 298), button='left')  
                ctrl_name = '정형 쿼리'
                
            elif group2 == 'time':
                time.sleep(0.5)
                mo.click(coords=(152, 315), button='left')  
                ctrl_name = '시간'
                
            elif group2 == 'command':
                time.sleep(0.5)
                ctrl_name = '명령어'
                
            elif group2 == 'event':
                time.sleep(0.5)
                ctrl_name = '이벤트'
            
            elif group2 == 'alert':
                time.sleep(0.5)
                ctrl_name = '경고'

        if group2=='service' or group2=='dbms' and group3=='access' or group2=='dbms' and group3=='auth' :
            element = shared.PARENT_WINDOW.child_window(title=ctrl_name, control_type='TreeItem', found_index=0)  # name(title)

        elif group2=='ftp' and group3=='access' or group2=='ftp' and group3=='auth' or group2=='sa':
            element = shared.PARENT_WINDOW.child_window(title=ctrl_name, control_type='TreeItem', found_index=1)  # name(title)

        elif group2=='terminal' and group3=='access' or group2=='terminal' and group3=='auth':
            element = shared.PARENT_WINDOW.child_window(title=ctrl_name, control_type='TreeItem', found_index=2)  # name(title)

        else:
            element = shared.PARENT_WINDOW.child_window(title=ctrl_name, control_type='TreeItem')  # name(title)

        try:
            if element.exists(timeout=3):
                rect = element.rectangle()
                center_x, center_y = rect.left + (rect.right - rect.left) // 2, rect.top + (rect.bottom - rect.top) // 2
                mo.click(coords=(center_x, center_y), button='left')
            else:
                print("Element not found within the timeout")

        except Exception as e:
            print(f"Error interacting with element: {e}")

            
    # 탭 필터링 버튼 없으면 누르기
    time.sleep(2)
    # time.sleep(1)
    
    try:
        # 현재 창에서 요소를 찾는다. 2024-09-03 수정
        ctrl_name = '자동 숨기기'
        my_utils.focus_window()
        element = shared.PARENT_WINDOW.child_window(title=ctrl_name, control_type='Button')  # auto_id
        
        if element.exists(timeout=2):
            print(f"필터링 고정 확인")
            
        else:
            print(f"필터링 고정 수행 ")
            if group1 == 'policy':
                mo.click(coords=(380, 230), button='left')    # 하드코딩 필요 

                element = shared.PARENT_WINDOW.child_window(title=ctrl_name, control_type='Button')
                if element.exists(timeout=2):
                    rect = element.rectangle()
                    center_x, center_y = rect.left + (rect.right - rect.left) // 2, rect.top + (rect.bottom - rect.top) // 2
                    mo.click(coords=(center_x, center_y), button='left')
                    
            elif group1 == 'object':
                if group2 == 'service':
                    mo.click(coords=(380, 235), button='left')
                    element = shared.PARENT_WINDOW.child_window(title=ctrl_name, control_type='Button')  # name(title)
        
                    if element.exists(timeout=2):
                        rect = element.rectangle()
                        center_x, center_y = rect.left + (rect.right - rect.left) // 2, rect.top + (rect.bottom - rect.top) // 2
                        mo.click(coords=(center_x, center_y), button='left')           
                
                else:
                    mo.click(coords=(380, 150), button='left')
                    element = shared.PARENT_WINDOW.child_window(title=ctrl_name, control_type='Button')  # name(title)
                    
                    if element.exists(timeout=2):
                        rect = element.rectangle()
                        center_x, center_y = rect.left + (rect.right - rect.left) // 2, rect.top + (rect.bottom - rect.top) // 2
                        mo.click(coords=(center_x, center_y), button='left')     
        
    except Exception as e:
        log_handler.log_exception(logger, e)
        # print(f"child_window 함수 실행 실패 : {e}")
        
def click_element_by_id_and_rectangle(app, automation_id, bounding_rect):
    # # 윈도우 핸들을 기반으로 윈도우 객체를 가져옵니다.
    # window = app.window(title=window_title)
    
    # 윈도우 제목에 "Enterprise Manager"가 포함된 첫 번째 윈도우를 찾습니다.
    window = app.window(title_re=".*Enterprise Manager.*")
    # window.print_control_identifiers()
    
    # AutomationId를 사용하여 모든 해당 요소를 찾습니다.
    elements = window.children(control_type="Pane", auto_id=automation_id)

    # 각 요소의 BoundingRectangle을 확인하여 클릭할 요소를 찾습니다.
    for element in elements:
        if element.exists(timeout=10):
            rect = element.rectangle()
            if (rect.left == bounding_rect['left'] and
                rect.top == bounding_rect['top'] and
                rect.right == bounding_rect['right'] and
                rect.bottom == bounding_rect['bottom']):
                # BoundingRectangle이 일치하는 요소의 중심 좌표를 계산합니다.
                center_x = rect.left + (rect.right - rect.left) // 2
                center_y = rect.top + (rect.bottom - rect.top) // 2

                # 해당 좌표를 클릭합니다.
                mo.click(coords=(center_x, center_y))
                print(f"Clicked on element with AutomationId: {automation_id} at ({center_x}, {center_y})")
                return  # 클릭 후 함수 종료

    print(f"No element with AutomationId: {automation_id} and the specified bounding rectangle found.")


def print_message(message, total_length=50, filler_char='='):
    """
    Print a message centered within a fixed-length line.

    :param message: The message to print.
    :param total_length: The total length of the line (default is 50 characters).
    :param filler_char: The character used to fill the line (default is '=').
    """
    # Calculate the padding needed on each side
    padding_length = (total_length - len(message) - 2) // 2  # Subtract 2 for the spaces around the message
    padding = filler_char * padding_length

    # Combine padding, message, and spaces
    formatted_message = f"{padding} {message} {padding}"

    # If the total length is odd, add one more filler character to the end
    if len(formatted_message) < total_length:
        formatted_message += filler_char

    print("\n"+formatted_message)
    
    
def mark_result(test_results, times):
    print_message("Mark Result")
    file_path = my_utils.get_xlsx()
    kill_excel_processes()
    time.sleep(3)   # 프로세스 안전 종료하기 위함
    try:
        wb = load_workbook(file_path)
        
    except Exception as e:
        print(f"Error Loading WorkBook: {e}, 파일이 손상되었을 수 있습니다. 임시파일을 생성합니다.")
        # 임시 파일 경로 설정 
        file_header, file_ext = os.path.splitext(file_path)
        temp_file = file_header + '_temp' + '.xlsx'
        
        # 임시 파일 생성
        wb = Workbook()
        ws = wb.active
        ws.title = shared.SHEET_NAME
        # ws.append(['id', 'product', 'summary', 'type', 'version', 'group', 'precondition', 'macro list', 'macro steps', 'val type', 'val data', 'expected result'])
        ws.append(shared.SHEET_HEADER)
        wb.save(temp_file)
        wb = load_workbook(temp_file)  # 새로 생성한 임시 파일을 다시 엽니다.
        file_path = temp_file  # 임시 파일 경로를 file_path로 업데이트하여 이후 로직에서 사용
    
    original_name = wb.sheetnames[0]
    original_sheet = wb[original_name]
    new_sheet = wb.copy_worksheet(original_sheet)
    new_sheet.title = f'{original_name}_Result'

    # Add the 'result' column with header and fill in the test results
    new_sheet.cell(row=1, column=new_sheet.max_column + 1, value='result')  # Header 추가

    # result column index
    result_col = new_sheet.max_column

    # Define the fills for pass (green) and fail (red)
    pass_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    fail_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    
    # for i in range(1, len(test_results) + 1):
    
    # row_idx = 2
    for k, v in test_results.items():
        
        result_text = 'Pass' if v == 1 else 'Fail'
        # cell = new_sheet.cell(row=row_idx, column=result_col, value=result_text)
        cell = new_sheet.cell(row=k+1, column=result_col, value=result_text)
        
        # Apply the fill color based on the result
        if test_results[k] == 1:
            cell.fill = pass_fill
        else:
            cell.fill = fail_fill
            
        # row_idx+=1
    
    # Write the start_time, end_time, and elapsed_time to the specified cells
    start_time, end_time = times
    elapsed_time = round(end_time - start_time,2)

    new_sheet.cell(row=2, column=17, value="start_time")
    new_sheet.cell(row=2, column=18, value=time.ctime(start_time))
    
    new_sheet.cell(row=3, column=17, value="end_time")
    new_sheet.cell(row=3, column=18, value=time.ctime(end_time))
    
    new_sheet.cell(row=4, column=17, value="elapsed_time")
    new_sheet.cell(row=4, column=18, value=f"{elapsed_time:.2f} seconds")

    # # Save the modified workbook
    # wb.save(file_path)
    # print("Results added and saved successfully.")
    
    # Save the modified workbook
    try:
        wb.save(file_path)
        print(f"Saved {file_path} Successfully")
        
    except Exception as e:
        print(f"Error saving workbook: {e}")


def kill_excel_processes():
    print_message("Kill Excel Process")

    # 현재 실행 중인 프로세스 목록을 가져옵니다.
    for proc in psutil.process_iter(['pid', 'name']):
        try:
            # 프로세스 이름이 'EXCEL.EXE'인 경우 프로세스를 종료합니다.
            if proc.info['name'].lower() == 'excel.exe':
                print(f"Killing process: {proc.info['name']} (PID: {proc.info['pid']})")
                proc.terminate()  # 프로세스 종료
                proc.wait(timeout=5)  # 프로세스가 종료될 때까지 기다림
        except (psutil.NoSuchProcess, psutil.AccessDenied, psutil.ZombieProcess):
            pass

# 2024-10-07 수정 - Excel 비정상 복구되는 현상 때문에 안정적으로 Close되는 코드로 변경
# def kill_excel_processes(s):
#     print_message("Closing Excel safely")
    # excel = win32com.client.Dispatch("Excel.Application")
    # excel.Workbooks.Close()  # 모든 열려있는 Excel 파일 닫기
    # excel.Quit()  # Excel 애플리케이션 종료

    # try:
    #     book = xw.Book(xlsx)
    #     book.close()

    # except Exception as e:
    #     print(e)

    

# 파일 내보내기로 download에 생성된 파일들 삭제
def delete_base_files():
    print_message("Delete Base Files")
    # 현재 사용자 홈 디렉토리를 가져옵니다.
    user_home = os.path.expanduser("~")

    # 사용자 홈 디렉토리를 기반으로 다운로드 경로를 생성합니다.
    downloads_path = os.path.join(user_home, "Downloads")
    
    # "_object"이 포함된 모든 CSV 파일을 찾습니다.
    # search_pattern = os.path.join(downloads_path, "*_object*.csv") # 20240930 주석처리
    search_pattern = os.path.join(downloads_path, "*.csv")
    files_to_delete = glob.glob(search_pattern)
    
    # 파일 삭제
    for file_path in files_to_delete:
        try:
            os.remove(file_path)
            print(f"Deleted: {file_path}")
        except OSError as e:
            print(f"Error deleting file {file_path}: {e}")


def create_base_dir():
    print_message("Create Base Dir")
    json_data = read_config(shared.CONFIG)
    
    # 기본 경로
    base_path = json_data['log']['base_path']

    # 기본 경로 생성
    if not os.path.exists(base_path):
        os.makedirs(base_path)
        print(f"Created directory: {base_path}")
    else:
        print(f"Already Exists: {base_path}")

    # 하위 디렉토리 생성
    for dir_name in json_data['log']['dir_name']:
        dir_path = os.path.join(base_path, dir_name)
        if not os.path.exists(dir_path):
            os.makedirs(dir_path)
            print(f"Created directory: {dir_path}")
        else:
            print(f"Already Exists: {dir_path}")

# 가장 최근에 생성된 파일 json file return
def get_json(file_name):
    file_name = file_name + '.json'
    json_dir = shared.JSON_DIR
    all_files = os.listdir(json_dir)
    # print(all_files)
    if file_name in all_files:
        print(f"Get File : {file_name}")
        return os.path.join(json_dir, file_name)
    else:
        print(f"File Not Exists ...")
        sys.exit(1)
        

# 가장 최근에 생성된 파일 xlsx file return
def get_xlsx():
    json_data = read_config(shared.CONFIG)
    base_path = json_data['log']['base_path']
    xlsxs = json_data['log']['dir_name'][3]
    
    xlsx_dir = os.path.join(base_path, xlsxs)
    all_files = os.listdir(xlsx_dir)
    # xlsx_files = [file for file in all_files if os.path.splitext(file)[1] == ".xlsx"]
    # '~$'로 시작하는 파일을 제외하고 .xlsx 확장자를 가진 파일만 선택
    xlsx_files = [file for file in all_files if file.endswith(".xlsx") and not file.startswith("~$")]
    xlsx_files.sort()
    # print(xlsx_files)
    
    if xlsx_files:
        shared.GSHEET = os.path.join(xlsx_dir, xlsx_files[-1])
        print(f"Load Existing Sheet : {shared.GSHEET}")
        # return shared.GSHEET
    else:
        sheet = spreadsheet.SpreadSheets()
        shared.GSHEET = sheet.save_sheet()
        print(f"Download from Google Sheet : {shared.GSHEET}")
        
    return shared.GSHEET

# config.yaml 파일 읽기 (환경설정 파일)
def read_config(file_path):
    with open(file_path, 'r') as file:
        config = yaml.safe_load(file)
    return config

# JSON 파일 읽기 
def read_json(file_path):
    try:
        with open(file_path, 'r', encoding='utf-8') as file:
            data = json.load(file)
        return data
    except FileNotFoundError:
        print(f"Error: The file '{file_path}' was not found.")
    except json.JSONDecodeError as e:
        print(f"Error: Failed to decode JSON. There might be a syntax error in the file '{file_path}'.")
        print(f"Details: {e}")
    except Exception as e:
        print(f"An unexpected error occurred while reading the file '{file_path}': {e}")
    return None  

# JSON 파일 쓰기
def write_json(file_path, data):
    with open(file_path, 'w', encoding='utf-8') as file:
        json.dump(data, file, ensure_ascii=False, indent=4)
    print(f"Save {file_path} Successfully")


def set_foreground():
    import win32gui
    import win32con
    import win32api
    import win32process
    import time
    
    shared.PID = get_process_id()
    
    # PID 기반으로 창의 핸들(hwnd)을 가져옴
    hwnd = get_hwnd_from_pid(shared.PID)

    try:
        # 창이 최소화된 상태일 경우 복원
        if win32gui.IsIconic(hwnd):
            win32gui.ShowWindow(hwnd, win32con.SW_RESTORE)

        # 현재 스레드와 창의 스레드를 연결하여 입력을 전환
        thread_id = win32api.GetCurrentThreadId()
        fg_thread_id = win32process.GetWindowThreadProcessId(win32gui.GetForegroundWindow())[0]
        if thread_id != fg_thread_id:
            win32process.AttachThreadInput(fg_thread_id, thread_id, True)

        # 창을 포그라운드로 설정
        win32gui.BringWindowToTop(hwnd)
        win32gui.SetForegroundWindow(hwnd)

        # 일정 시간 대기 후 창을 맨 앞으로 이동
        time.sleep(0.1)  # 지연 시간 추가
        win32gui.SetWindowPos(hwnd, win32con.HWND_TOPMOST, 0, 0, 0, 0,
                              win32con.SWP_NOMOVE | win32con.SWP_NOSIZE)
        win32gui.SetWindowPos(hwnd, win32con.HWND_NOTOPMOST, 0, 0, 0, 0,
                              win32con.SWP_NOMOVE | win32con.SWP_NOSIZE)

        # 입력 연결 해제
        if thread_id != fg_thread_id:
            win32process.AttachThreadInput(fg_thread_id, thread_id, False)

        print(f"SetForegroundWindow 성공")

    except Exception as e:
        print(f"SetForegroundWindow 실패: {e}")


# 화면 최대화
def maximize_window():
    # shared.PID = 1640
    shared.PID = get_process_id()
    hwnd = get_hwnd_from_pid(shared.PID)
    if hwnd is None:
        print("No window found with the given PID.")
        return
    curr_window = win32gui.GetWindowText(hwnd)
    
    # cond1 = shared.PROCESS_NAME == 'Enterprise Manager.exe'   # 주석처리 2024-08-28 16:24
    cond2 = 'login' in curr_window.lower()
    cond3 = '알림' in curr_window
    
    # if cond1 and not cond2 :  # 주석처리 2024-08-28 16:24
    if not cond2 and not cond3:
        win32gui.ShowWindow(hwnd, win32con.SW_MINIMIZE)
        win32gui.ShowWindow(hwnd, win32con.SW_MAXIMIZE)
        
    else:
        print(f"창을 최대화 하지 않음")
    

def focus_window():
    shared.PID = get_process_id()
    # set_foreground() 
    hwnd = get_hwnd_from_pid(shared.PID)
    curr_window = win32gui.GetWindowText(hwnd)
    
    try:
        app = Application(backend='uia').connect(handle=hwnd)
        
        # 현재 프로세스에서 열려 있는 모든 창의 핸들 검색
        window_handles = find_windows(process=shared.PID, title_re=".*")
        for handle in window_handles:
            window = app.window(handle=handle)
            window_title = window.window_text()
            
            if curr_window in window_title:
                print(f"Matching window found: {window_title}")
                shared.PARENT_WINDOW = window
                break
        else:
            print("No matching window found.")
            return
        # time.sleep(2)
        time.sleep(1) # 2024-09-30 수정

    except ElementNotFoundError:
        print(f"No window found with title containing '{curr_window}'")
        
    except Exception as e:
        print(f"An unexpected error occurred: {e}")
        
    

# 좌표 클릭 
def click_with_coordinates(parent_window, ctrl_id=None, ctrl_name=None, ctrl_type=None, coords=None):
    elements = None
    
    try:
        elements = findwindows.find_elements(
            parent=parent_window.element_info,
            auto_id=ctrl_id,
            title=ctrl_name,
            control_type=ctrl_type,
            top_level_only=False,
            enabled_only=True,
            visible_only=True
        )
        for element in elements:
            element_info = UIAElementInfo(element.element)
            rect = element_info.rectangle
            x1, y1 = rect.left, rect.top
            width, height = rect.right - x1, rect.bottom - y1
            x2, y2 = x1 + width, y1 + height  
            saved_x, saved_y = coords[0], coords[1]
            if x1 <= saved_x < x2 and y1 <= saved_y < y2:
                return rect
    except:
        return elements

# 프로세스 Kill
def kill_process_by_pname(process_name, timeout=1):
    for proc in psutil.process_iter(['pid', 'name']):
        try:
            if proc.info['name'] == process_name:
                print(f"Terminating process {proc.info['name']} with PID {proc.info['pid']}")
                proc.terminate()
                try:
                    proc.wait(timeout=timeout)
                    print(f"Process {proc.info['name']} with PID {proc.info['pid']} terminated successfully")
                    time.sleep(1)
                    
                except psutil.TimeoutExpired:
                    print(f"Process {proc.info['name']} with PID {proc.info['pid']} did not terminate in {timeout} seconds.")
        except (psutil.NoSuchProcess, psutil.AccessDenied, psutil.ZombieProcess) as e:
            print(f"Could not terminate process {proc.info['name']} with PID {proc.info['pid']}: {e}")

def close_alert_windows(app, alert_title='알림'):
    for _ in range(2):  # 알림창이 2번 생성되는 경우
        try:
            # print(f"{i+1} 회차 esc 2번 press 실행")   
            # 알림창이 나타날 때까지 기다림 (최대 5초)
            # alert_window = app.window(best_match=alert_title)
            
            alert_window = app.window(title_re=f".*{alert_title}.*")
            print(f"alert_window : {alert_window}")
            
            alert_window.wait('exists', timeout=5)
            alert_window.set_focus()

            # ESC 키를 두 번 눌러 알림창을 닫기
            alert_window.type_keys("{ESC}")
            time.sleep(0.5)
            alert_window.type_keys("{ESC}")
            
        except Exception as e:
            print(f"{alert_title} 창을 닫는 중 오류 발생: {e}")
            
            
# 매니저 로그인
def login_manager(pid):
    data = read_config(shared.CONFIG)
    server_ip = data['db_connection']['db_ip']
    mgr_id = data['start']['user_name']
    mgr_passwd = data['start']['user_pw']
    
    app = Application(backend='uia').connect(process=pid)
    print("프로세스 ID:", app.process)
    print("윈도우 목록:", app.windows())
    if app.windows() == []:
        return False

    # 메인 윈도우의 핸들 가져오기
    app.top_window()
    
    mo.move(coords=(960, 600))
    
    parent = app.window(title='DBSAFER Enterprise Manager Login')
    parent.wait("ready", timeout=5)

    ip_field = parent.window(auto_id="1001", control_type="Edit").wrapper_object()
    ip_field.set_edit_text(u'')
    ip_field.type_keys(server_ip)
    
    id_field = parent.window(auto_id="1305", control_type="Edit").wrapper_object()
    id_field.set_edit_text(u'')
    id_field.type_keys(mgr_id)

    pwd_field = parent.window(auto_id="1301", control_type="Edit").wrapper_object()
    pwd_field.set_text(mgr_passwd)
    pwd_field.type_keys("{ENTER}")
    time.sleep(5)   # 알림창 팝업까지 대기 시간 
    close_alert_windows(app)


# pid로 hwnd값 가져오기 
def get_hwnd_from_pid(pid):
    hwnds = []
    def callback(hwnd, hwnds):
        if win32gui.IsWindowVisible(hwnd) and win32gui.IsWindowEnabled(hwnd):
            _, found_pid = win32process.GetWindowThreadProcessId(hwnd)
            if found_pid == pid:
                hwnds.append(hwnd)
        return True

    win32gui.EnumWindows(callback, hwnds)
    if hwnds:
        return hwnds[0]
    return None

# PROCESS_NAME의 pid 가져오기 
def get_process_id():
    pid = None
    for proc in psutil.process_iter(attrs=['pid', 'name']):
        if proc.info['name'] == shared.PROCESS_NAME:
            pid = proc.info['pid']
            break
    return pid


    
if __name__=='__main__':
    # xlsx = get_xlsx()
    # kill_excel_processes(xlsx)

    # First, check for any background Excel processes and kill them if needed
    # check_and_kill_existing_excel()

    # Call the function to gracefully close Excel
    # close_excel_gracefully()

    # maximize_window()
    # focus_window()
    # release_services()

    # policys = ['policy_dbms_access', 'policy_dbms_auth', 'policy_dbms_global', 'policy_ftp_access', 'policy_ftp_auth', 'policy_terminal_access', 'policy_terminal_auth',
    #            'policy_sa_dbms', 'policy_sa_ftp', 'policy_sa_terminal']
    # for obj in policys[:]:
    #     click_into_group(obj)

    # click_into_group('policy_ftp_access')


    # objects = ['object_service_dbms', 'object_service_ftp', 'object_service_terminal', 'object_service_transparent', 'object_instance', 'object_ip', 'object_access', 'object_application', 'object_cert', 'object_datamask', 'object_tablecolumn', 'object_formalquery', 'object_time', 'object_command', 'object_alert']
    # for obj in objects:
    #     click_into_group(obj)
        
    # test_results = {371: 1, 372: 1, 373: 1, 374: 1, 375: 1, 376: 1, 377: 1, 378: 1, 379: 1, 380: 1, 381: 1, 382: 1, 383: 1, 384: 1, 385: 1, 386: 1, 387: 1, 388: 1, 389: 1, 390: 1, 391: 1, 392: 1, 393: 1, 394: 1, 395: 1, 396: 1, 397: 1, 398: 1, 399: 1, 400: 1, 401: 1, 402: 1, 403: 1, 404: 1, 405: 1, 406: 1}
    # mark_result(test_results, [900,8])
    pass 